import asyncio
import json
import os
import re
import shutil
import tempfile
import time as time_module
import uuid
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openai import OpenAI

from config import (
    DEEPSEEK_API_KEY,
    DEEPSEEK_MODEL,
    EXCEL_AI_PROVIDER,
    EXCEL_OPENAI_MODEL,
    OPENAI_API_KEY,
    logger,
)
from database import add_to_conversation
from group_utils import get_dialog_key

DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"
openai_excel_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

SUPPORTED_EXCEL_EXTENSIONS = (".xlsx",)
UNSUPPORTED_EXCEL_EXTENSIONS = (".xls", ".xlsm", ".xlsb", ".ods")

HEADER_FILL = "D9EAF7"
LIGHT_FILL = "F3F6FA"
BORDER_COLOR = "D9E2EC"


class ExcelProcessingError(Exception):
    pass


def is_excel_file(file_name: str) -> bool:
    return (file_name or "").lower().endswith(SUPPORTED_EXCEL_EXTENSIONS + UNSUPPORTED_EXCEL_EXTENSIONS)


def is_supported_excel_file(file_name: str) -> bool:
    return (file_name or "").lower().endswith(SUPPORTED_EXCEL_EXTENSIONS)


def is_create_excel_request(text: str) -> bool:
    if not text:
        return False
    lowered = text.lower()
    create_words = ("создай", "сделай", "сформируй", "подготовь", "собери")
    has_create_word = any(word in lowered for word in create_words)
    has_excel_file_word = any(word in lowered for word in ("excel", "xlsx", "эксель", "ексель", "файл"))
    if any(phrase in lowered for phrase in ("сделай вывод", "сделай анализ", "подготовь вывод")):
        return False
    # Ловим фразы вида «создай таблицу...» или «сделай таблицу...»,
    # но не перехватываем «сделай выводы по таблице».
    has_create_table_phrase = bool(
        re.search(r"(создай|сформируй|подготовь|собери).{0,50}таблиц", lowered)
        or re.search(r"сделай\s+таблиц", lowered)
    )
    return has_create_word and (has_excel_file_word or has_create_table_phrase)


def looks_like_excel_edit_request(text: str) -> bool:
    if not text:
        return False
    lowered = text.lower()
    edit_patterns = [
        "измени", "исправь", "поправь", "отредактируй", "добавь", "удали", "замени",
        "переименуй", "заполни", "посчитай", "рассчитай", "вставь", "оформи", "отформатируй",
        "сделай новый лист", "добавь лист", "добавь столбец", "новый столбец", "новую колонку",
        "создай лист", "создай столбец", "создай колонку", "формулу", "формулы", "итоговый лист",
        "свод", "сводную", "отсортируй", "поставь фильтр", "сделай фильтр",
    ]
    analysis_patterns = ["проанализируй", "посмотри", "что", "почему", "найди", "проверь", "скажи"]
    if any(pattern in lowered for pattern in edit_patterns):
        return True
    # «сделай ... в файле» обычно ожидает измененный файл, а не только ответ.
    if "сделай" in lowered and any(word in lowered for word in ("лист", "столбец", "колонк", "итог", "таблиц", "формул", "файл")):
        return True
    return False


def _safe_filename(name: str, default: str = "excel_result.xlsx") -> str:
    name = (name or default).strip().replace("\\", "_").replace("/", "_")
    name = re.sub(r"[^\w\-.а-яА-ЯёЁ ]+", "_", name, flags=re.UNICODE).strip(" ._")
    if not name:
        name = default
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return name[:120]


def _excel_cache_dir() -> str:
    """Папка для последнего Excel-файла пользователя.

    На Railway можно задать EXCEL_STORAGE_DIR=/data/excel_cache.
    Если не задано, но есть DB_FILE=/data/bot_data.db, файлы кладутся рядом с базой в /data/excel_cache.
    Иначе используется временная папка контейнера.
    """
    explicit = os.getenv("EXCEL_STORAGE_DIR")
    if explicit:
        base = explicit
    else:
        db_file = os.getenv("DB_FILE")
        if db_file:
            base = os.path.join(os.path.dirname(os.path.abspath(db_file)), "excel_cache")
        elif os.getenv("RAILWAY_VOLUME_MOUNT_PATH"):
            base = os.path.join(os.getenv("RAILWAY_VOLUME_MOUNT_PATH"), "excel_cache")
        else:
            base = os.path.join(tempfile.gettempdir(), "mitra_excel_cache")
    os.makedirs(base, exist_ok=True)
    return base


def remember_excel_file(context, chat_id: int, source_path: str, file_name: str, dialog_key: Optional[str] = None) -> Dict[str, Any]:
    """Сохраняет присланный/созданный Excel, чтобы следующий текст пользователя мог его изменить."""
    dialog_key = str(dialog_key or chat_id)
    excel_files = context.user_data.setdefault("excel_files", {}) if hasattr(context, "user_data") else {}
    old = excel_files.get(dialog_key)
    if old and old.get("path") and os.path.exists(old["path"]):
        try:
            os.unlink(old["path"])
        except OSError:
            pass

    safe_name = _safe_filename(file_name or "table.xlsx")
    stored_name = f"chat_{chat_id}_{uuid.uuid4().hex[:8]}_{safe_name}"
    stored_path = os.path.join(_excel_cache_dir(), stored_name)
    shutil.copy2(source_path, stored_path)

    info = {
        "path": stored_path,
        "file_name": safe_name,
        "saved_at": time_module.time(),
    }
    excel_files[dialog_key] = info
    context.user_data["awaiting_excel_request_by_dialog"] = {
        **context.user_data.get("awaiting_excel_request_by_dialog", {}),
        dialog_key: True,
    }
    return info


def _get_recent_excel_context(context, dialog_key: str, max_age_seconds: int = 24 * 60 * 60) -> Optional[Dict[str, Any]]:
    dialog_key = str(dialog_key)
    excel_files = context.user_data.get("excel_files", {}) if hasattr(context, "user_data") else {}
    info = excel_files.get(dialog_key)
    if not info:
        return None
    path = info.get("path")
    if not path or not os.path.exists(path):
        excel_files.pop(dialog_key, None)
        context.user_data.get("awaiting_excel_request_by_dialog", {}).pop(dialog_key, None)
        return None
    if time_module.time() - float(info.get("saved_at") or 0) > max_age_seconds:
        excel_files.pop(dialog_key, None)
        context.user_data.get("awaiting_excel_request_by_dialog", {}).pop(dialog_key, None)
        return None
    return info


def _is_excel_followup_request(context, text: str, dialog_key: str) -> bool:
    if not text:
        return False
    lowered = text.lower()
    awaiting = context.user_data.get("awaiting_excel_request_by_dialog", {}) if hasattr(context, "user_data") else {}
    if awaiting.get(str(dialog_key)):
        return True
    if looks_like_excel_edit_request(text):
        return True
    return any(word in lowered for word in ("excel", "xlsx", "эксель", "ексель", "таблиц", "лист", "столбец", "колонк", "ячейк", "формул"))


def _cell_value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value)


def _normalize_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return _cell_value_to_text(value)
    return value


def _trim_text(text: str, limit: int = 45000) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + "\n... [данные обрезаны из-за размера файла]"


def workbook_to_text(path: str, max_rows_per_sheet: int = 80, max_cols_per_sheet: int = 25) -> str:
    """Делает текстовое представление Excel для ИИ: размеры листов, заголовки, первые строки."""
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        raise ExcelProcessingError(f"openpyxl не смог прочитать книгу: {e}") from e
    parts: List[str] = []
    parts.append(f"Книга содержит листы: {', '.join(wb.sheetnames)}")

    for ws in wb.worksheets:
        parts.append(f"\n=== Лист: {ws.title} ===")
        parts.append(f"Размер: {ws.max_row} строк x {ws.max_column} столбцов")

        if ws.merged_cells.ranges:
            merged = ", ".join(str(rng) for rng in list(ws.merged_cells.ranges)[:20])
            parts.append(f"Объединенные ячейки: {merged}")

        formulas = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows_per_sheet), min_col=1, max_col=min(ws.max_column, max_cols_per_sheet)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{cell.coordinate}={cell.value[:120]}")
                    if len(formulas) >= 20:
                        break
            if len(formulas) >= 20:
                break
        if formulas:
            parts.append("Формулы: " + "; ".join(formulas))

        rows_limit = min(ws.max_row, max_rows_per_sheet)
        cols_limit = min(ws.max_column, max_cols_per_sheet)
        if rows_limit == 0 or cols_limit == 0:
            parts.append("Лист пустой")
            continue

        for row in ws.iter_rows(min_row=1, max_row=rows_limit, min_col=1, max_col=cols_limit):
            values = [_cell_value_to_text(cell.value) for cell in row]
            # Обрезаем длинные тексты в ячейках.
            values = [v if len(v) <= 220 else v[:220] + "..." for v in values]
            parts.append(f"R{row[0].row}: " + " | ".join(values))

        if ws.max_row > rows_limit:
            parts.append(f"... показаны первые {rows_limit} строк из {ws.max_row}")
        if ws.max_column > cols_limit:
            parts.append(f"... показаны первые {cols_limit} столбцов из {ws.max_column}")

    return _trim_text("\n".join(parts))


def _call_deepseek_for_excel(system_prompt: str, user_prompt: str, max_tokens: int = 2500, temperature: float = 0.1) -> str:
    if not DEEPSEEK_API_KEY:
        raise ExcelProcessingError("Не настроен DEEPSEEK_API_KEY")

    try:
        excel_token_limit = int(os.getenv("EXCEL_MAX_RESPONSE_TOKENS", "6000"))
    except ValueError:
        excel_token_limit = 6000

    data = {
        "model": DEEPSEEK_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_tokens": max(512, min(excel_token_limit, max_tokens)),
        "temperature": temperature,
    }
    headers = {
        "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
        "Content-Type": "application/json",
    }
    try:
        response = requests.post(DEEPSEEK_API_URL, headers=headers, json=data, timeout=120)
        response.raise_for_status()
        payload = response.json()
        choice = payload["choices"][0]
        content = (choice.get("message") or {}).get("content") or ""
        finish_reason = choice.get("finish_reason") or "unknown"
        if not content.strip():
            raise ExcelProcessingError(f"ИИ вернул пустой ответ (finish_reason={finish_reason})")
        return content
    except ExcelProcessingError:
        raise
    except requests.HTTPError as e:
        detail = (getattr(e.response, "text", "") or "")[:500]
        status = getattr(e.response, "status_code", "unknown")
        raise ExcelProcessingError(f"ИИ-сервис вернул ошибку HTTP {status}. {detail}") from e
    except requests.RequestException as e:
        raise ExcelProcessingError(f"ИИ-сервис временно недоступен или не ответил: {e}") from e
    except (KeyError, IndexError, ValueError) as e:
        raise ExcelProcessingError("ИИ-сервис вернул ответ в неожиданном формате") from e


def _call_openai_for_excel(system_prompt: str, user_prompt: str, max_tokens: int = 2500, temperature: float = 0.1) -> str:
    if not openai_excel_client:
        raise ExcelProcessingError("Не настроен OPENAI_API_KEY для обработки Excel")

    try:
        excel_token_limit = int(os.getenv("EXCEL_MAX_RESPONSE_TOKENS", "6000"))
    except ValueError:
        excel_token_limit = 6000

    try:
        response = openai_excel_client.chat.completions.create(
            model=EXCEL_OPENAI_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            max_tokens=max(512, min(excel_token_limit, max_tokens)),
            temperature=temperature,
        )
        choice = response.choices[0]
        content = choice.message.content or ""
        finish_reason = choice.finish_reason or "unknown"
        if not content.strip():
            raise ExcelProcessingError(f"OpenAI вернул пустой ответ (finish_reason={finish_reason})")
        return content
    except ExcelProcessingError:
        raise
    except Exception as e:
        raise ExcelProcessingError(f"OpenAI не смог обработать Excel-запрос: {e}") from e


def _call_ai_for_excel(system_prompt: str, user_prompt: str, max_tokens: int = 2500, temperature: float = 0.1) -> str:
    if EXCEL_AI_PROVIDER == "openai":
        return _call_openai_for_excel(system_prompt, user_prompt, max_tokens=max_tokens, temperature=temperature)
    if EXCEL_AI_PROVIDER == "auto" and openai_excel_client:
        try:
            return _call_openai_for_excel(system_prompt, user_prompt, max_tokens=max_tokens, temperature=temperature)
        except ExcelProcessingError as openai_error:
            logger.warning(f"OpenAI Excel fallback failed, trying DeepSeek: {openai_error}")
    return _call_deepseek_for_excel(system_prompt, user_prompt, max_tokens=max_tokens, temperature=temperature)


def _extract_json(text: str) -> Dict[str, Any]:
    if not text:
        raise ExcelProcessingError("ИИ вернул пустой ответ")
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r"```$", "", cleaned).strip()

    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(cleaned[start:end + 1])
        raise


def analyze_excel_with_ai(path: str, file_name: str, question: str, user_id: str) -> str:
    summary = workbook_to_text(path)
    question = (question or "Проанализируй Excel-файл и кратко опиши, что в нём есть.").strip()
    system_prompt = (
        "Ты — ассистент, который анализирует Excel-файлы. Отвечай на русском. "
        "Используй только данные из предоставленного текстового представления книги. "
        "Если данных недостаточно из-за обрезки, честно скажи об этом."
    )
    user_prompt = f"Файл: {file_name}\n\nДанные Excel:\n{summary}\n\nВопрос пользователя:\n{question}"
    answer = _call_ai_for_excel(system_prompt, user_prompt, max_tokens=3500, temperature=0.2)
    add_to_conversation(user_id, "user", f"[Excel {file_name}] {question}\n{summary[:4000]}")
    add_to_conversation(user_id, "assistant", answer)
    return answer


def _first_sheet_name(path: str) -> str:
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return wb.sheetnames[0]
    finally:
        wb.close()


def build_excel_patch_with_ai(path: str, file_name: str, request_text: str) -> Dict[str, Any]:
    summary = workbook_to_text(path, max_rows_per_sheet=120, max_cols_per_sheet=35)
    first_sheet = _first_sheet_name(path)
    system_prompt = """
Ты преобразуешь просьбу пользователя о правке Excel в безопасный JSON-план.
Отвечай ТОЛЬКО валидным JSON без markdown.
Не придумывай данные, которых нет в таблице. Если просьба неоднозначная — верни need_clarification=true.

Доступные действия:
1) add_sheet: {"type":"add_sheet","sheet":"Имя"}
2) rename_sheet: {"type":"rename_sheet","old_sheet":"Старое","new_sheet":"Новое"}
3) set_cell: {"type":"set_cell","sheet":"Лист","cell":"A1","value":"текст или число"}
4) set_formula: {"type":"set_formula","sheet":"Лист","cell":"E2","formula":"=C2*D2"}
5) append_row: {"type":"append_row","sheet":"Лист","values":["...",123]}
6) add_column_by_condition: добавляет столбец и заполняет строки 2..последняя строка по условию.
   Формат: {"type":"add_column_by_condition","sheet":"Лист","header":"Комментарий","condition_header":"Прогресс","operator":"<","value":50,"true_value":"Отстаёт","false_value":""}
   operator: <, <=, >, >=, ==, !=, contains, empty, not_empty
7) add_formula_column: добавляет столбец с формулами.
   Формат: {"type":"add_formula_column","sheet":"Лист","header":"Итого","formula_template":"=C{row}*D{row}","number_format":"0.00"}
8) create_summary_sheet: создает лист со сводкой по количеству строк по выбранному заголовку.
   Формат: {"type":"create_summary_sheet","source_sheet":"Лист","sheet":"Итоги","group_by_header":"Ответственный","sum_header":"Сумма"}
   sum_header можно не указывать, тогда будет только количество.
9) style_header: {"type":"style_header","sheet":"Лист","row":1}
10) auto_width: {"type":"auto_width","sheet":"Лист"}
11) freeze_panes: {"type":"freeze_panes","sheet":"Лист","cell":"A2"}
12) make_table: {"type":"make_table","sheet":"Лист","name":"Table1"}

Верни объект:
{
  "need_clarification": false,
  "message": "короткое описание, что будет изменено",
  "actions": [ ... ]
}

Правила:
- Если пользователь просит просто проанализировать/объяснить, а не изменить файл — верни actions: [] и message с объяснением.
- Если имя листа не указано, используй первый лист из книги.
- Для массовых операций по всем строкам используй add_column_by_condition или add_formula_column, а не set_cell для каждой строки.
- Не используй опасные внешние ссылки и макросы.
""".strip()

    user_prompt = (
        f"Файл: {file_name}\n"
        f"Первый лист по умолчанию: {first_sheet}\n\n"
        f"Данные Excel:\n{summary}\n\n"
        f"Просьба пользователя:\n{request_text}"
    )
    raw = _call_ai_for_excel(system_prompt, user_prompt, max_tokens=3500, temperature=0.0)
    patch = _extract_json(raw)
    if not isinstance(patch, dict):
        raise ExcelProcessingError("ИИ вернул не объект JSON")
    patch.setdefault("need_clarification", False)
    patch.setdefault("message", "Готовлю изменения в Excel-файле.")
    patch.setdefault("actions", [])
    if not isinstance(patch.get("actions"), list):
        raise ExcelProcessingError("Поле actions должно быть списком")
    return patch


def _get_ws(wb, sheet_name: Optional[str] = None):
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    if sheet_name and sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name[:31])
        return ws
    return wb.worksheets[0]


def _find_header_column(ws, header: str, header_row: int = 1) -> Optional[int]:
    target = (header or "").strip().lower()
    if not target:
        return None
    for cell in ws[header_row]:
        value = str(cell.value or "").strip().lower()
        if value == target:
            return cell.column
    for cell in ws[header_row]:
        value = str(cell.value or "").strip().lower()
        if target in value or value in target:
            return cell.column
    return None


def _last_used_column(ws) -> int:
    return max(ws.max_column or 1, 1)


def _as_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _compare(left: Any, operator: str, right: Any) -> bool:
    operator = (operator or "==").strip().lower()
    if operator in ("empty", "is_empty"):
        return left is None or str(left).strip() == ""
    if operator in ("not_empty", "not empty"):
        return not (left is None or str(left).strip() == "")
    if operator == "contains":
        return str(right).lower() in str(left or "").lower()

    left_num = _as_number(left)
    right_num = _as_number(right)
    if left_num is not None and right_num is not None:
        a, b = left_num, right_num
    else:
        a, b = str(left or "").strip().lower(), str(right or "").strip().lower()

    if operator == "<":
        return a < b
    if operator == "<=":
        return a <= b
    if operator == ">":
        return a > b
    if operator == ">=":
        return a >= b
    if operator in ("!=", "<>"):
        return a != b
    return a == b


def _style_header_row(ws, row: int = 1) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_width(ws, max_width: int = 42) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for cell in ws[letter]:
            value = _cell_value_to_text(cell.value)
            if value:
                width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def _add_excel_table(ws, name: str) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    # У таблицы должны быть заголовки.
    if not any(ws.cell(1, c).value for c in range(1, ws.max_column + 1)):
        return
    clean_name = re.sub(r"[^A-Za-z0-9_]", "_", name or f"Table_{ws.title}")
    if not clean_name or clean_name[0].isdigit():
        clean_name = "Table_" + clean_name
    clean_name = clean_name[:30]
    existing = {tbl.name for tbl in ws.tables.values()}
    base = clean_name
    idx = 1
    while clean_name in existing:
        clean_name = f"{base}_{idx}"[:30]
        idx += 1
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=clean_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def apply_excel_patch(input_path: str, output_path: str, patch: Dict[str, Any]) -> List[str]:
    wb = load_workbook(input_path)
    changes: List[str] = []

    for idx, action in enumerate(patch.get("actions", []), start=1):
        if not isinstance(action, dict):
            continue
        action_type = action.get("type")

        if action_type == "add_sheet":
            sheet = str(action.get("sheet") or "Новый лист")[:31]
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
                changes.append(f"Добавлен лист «{sheet}»")

        elif action_type == "rename_sheet":
            old = action.get("old_sheet")
            new = str(action.get("new_sheet") or "Лист")[:31]
            if old in wb.sheetnames and new not in wb.sheetnames:
                wb[old].title = new
                changes.append(f"Лист «{old}» переименован в «{new}»")

        elif action_type == "set_cell":
            ws = _get_ws(wb, action.get("sheet"))
            cell = action.get("cell") or "A1"
            ws[cell] = _normalize_value(action.get("value"))
            changes.append(f"Изменена ячейка {ws.title}!{cell}")

        elif action_type == "set_formula":
            ws = _get_ws(wb, action.get("sheet"))
            cell = action.get("cell") or "A1"
            formula = str(action.get("formula") or "")
            if formula and not formula.startswith("="):
                formula = "=" + formula
            ws[cell] = formula
            if action.get("number_format"):
                ws[cell].number_format = action["number_format"]
            changes.append(f"Добавлена формула в {ws.title}!{cell}")

        elif action_type == "append_row":
            ws = _get_ws(wb, action.get("sheet"))
            values = [_normalize_value(v) for v in action.get("values", [])]
            ws.append(values)
            changes.append(f"Добавлена строка на лист «{ws.title}»")

        elif action_type == "add_column_by_condition":
            ws = _get_ws(wb, action.get("sheet"))
            header = str(action.get("header") or "Новый столбец")
            condition_header = str(action.get("condition_header") or "")
            condition_col = _find_header_column(ws, condition_header)
            if not condition_col:
                changes.append(f"Не найден столбец условия «{condition_header}» на листе «{ws.title}»")
                continue
            new_col = _last_used_column(ws) + 1
            ws.cell(1, new_col).value = header
            operator = action.get("operator") or "=="
            value = action.get("value")
            true_value = _normalize_value(action.get("true_value", ""))
            false_value = _normalize_value(action.get("false_value", ""))
            for row in range(2, ws.max_row + 1):
                source_value = ws.cell(row, condition_col).value
                ws.cell(row, new_col).value = true_value if _compare(source_value, operator, value) else false_value
            changes.append(f"Добавлен столбец «{header}» по условию «{condition_header} {operator} {value}»")

        elif action_type == "add_formula_column":
            ws = _get_ws(wb, action.get("sheet"))
            header = str(action.get("header") or "Расчет")
            template = str(action.get("formula_template") or "")
            if not template:
                continue
            new_col = _last_used_column(ws) + 1
            ws.cell(1, new_col).value = header
            start_row = int(action.get("start_row") or 2)
            end_row = int(action.get("end_row") or ws.max_row)
            number_format = action.get("number_format")
            col_letter = get_column_letter(new_col)
            for row in range(start_row, end_row + 1):
                formula = template.format(row=row, col=col_letter)
                if not formula.startswith("="):
                    formula = "=" + formula
                cell = ws.cell(row, new_col)
                cell.value = formula
                if number_format:
                    cell.number_format = number_format
            changes.append(f"Добавлен расчетный столбец «{header}»")

        elif action_type == "create_summary_sheet":
            source_ws = _get_ws(wb, action.get("source_sheet"))
            group_header = str(action.get("group_by_header") or "")
            group_col = _find_header_column(source_ws, group_header)
            if not group_col:
                changes.append(f"Не найден столбец для сводки «{group_header}»")
                continue
            sum_header = action.get("sum_header")
            sum_col = _find_header_column(source_ws, str(sum_header)) if sum_header else None
            sheet_name = str(action.get("sheet") or "Итоги")[:31]
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            ws.append([group_header, "Количество"] + ([str(sum_header)] if sum_col else []))
            summary: Dict[str, Dict[str, Any]] = {}
            for row in range(2, source_ws.max_row + 1):
                key = _cell_value_to_text(source_ws.cell(row, group_col).value) or "(пусто)"
                summary.setdefault(key, {"count": 0, "sum": 0.0})
                summary[key]["count"] += 1
                if sum_col:
                    num = _as_number(source_ws.cell(row, sum_col).value)
                    if num is not None:
                        summary[key]["sum"] += num
            for key, data in sorted(summary.items(), key=lambda item: item[0]):
                ws.append([key, data["count"]] + ([data["sum"]] if sum_col else []))
            _style_header_row(ws, 1)
            _auto_width(ws)
            changes.append(f"Создан лист «{sheet_name}» со сводкой по «{group_header}»")

        elif action_type == "style_header":
            ws = _get_ws(wb, action.get("sheet"))
            _style_header_row(ws, int(action.get("row") or 1))
            changes.append(f"Оформлены заголовки на листе «{ws.title}»")

        elif action_type == "auto_width":
            ws = _get_ws(wb, action.get("sheet"))
            _auto_width(ws)
            changes.append(f"Подобрана ширина столбцов на листе «{ws.title}»")

        elif action_type == "freeze_panes":
            ws = _get_ws(wb, action.get("sheet"))
            ws.freeze_panes = action.get("cell") or "A2"
            changes.append(f"Закреплена область на листе «{ws.title}»")

        elif action_type == "make_table":
            ws = _get_ws(wb, action.get("sheet"))
            _add_excel_table(ws, action.get("name") or f"Table_{idx}")
            changes.append(f"Добавлен фильтр-таблица на листе «{ws.title}»")

    # Базовое аккуратное оформление, если были изменения.
    if patch.get("actions"):
        for ws in wb.worksheets:
            if ws.max_row >= 1 and ws.max_column >= 1:
                _style_header_row(ws, 1)
                _auto_width(ws)
                if ws.freeze_panes is None and ws.max_row > 1:
                    ws.freeze_panes = "A2"

    wb.save(output_path)
    return changes


def edit_excel_with_ai(input_path: str, file_name: str, request_text: str, user_id: str) -> Tuple[Optional[str], str, List[str]]:
    patch = build_excel_patch_with_ai(input_path, file_name, request_text)
    if patch.get("need_clarification"):
        return None, patch.get("message") or "Нужно уточнить, что именно изменить в Excel-файле.", []

    actions = patch.get("actions", [])
    if not actions:
        # Если ИИ решил, что это не правка, дадим обычный анализ.
        answer = analyze_excel_with_ai(input_path, file_name, request_text, user_id)
        return None, answer, []

    base = Path(file_name).stem or "excel"
    out_name = _safe_filename(f"{base}_edited_{uuid.uuid4().hex[:6]}.xlsx")
    output_path = os.path.join(tempfile.gettempdir(), out_name)
    changes = apply_excel_patch(input_path, output_path, patch)
    message = patch.get("message") or "Готово, внесла изменения в Excel-файл."

    add_to_conversation(user_id, "user", f"[Правка Excel {file_name}] {request_text}")
    add_to_conversation(user_id, "assistant", message + "\n" + "\n".join(changes))
    return output_path, message, changes


def _build_workbook_from_spec(spec: Dict[str, Any], output_path: str) -> None:
    wb = Workbook()
    # Удалим дефолтный лист после создания первого своего листа.
    default_ws = wb.active
    created_any = False

    sheets = spec.get("sheets") or []
    if not sheets:
        sheets = [{"name": "Лист1", "headers": ["Наименование", "Комментарий"], "rows": []}]

    for sheet_spec in sheets[:10]:
        name = str(sheet_spec.get("name") or "Лист")[:31]
        ws = wb.create_sheet(name)
        created_any = True
        headers = sheet_spec.get("headers") or []
        rows = sheet_spec.get("rows") or []
        if headers:
            ws.append([_normalize_value(v) for v in headers])
        for row in rows[:500]:
            ws.append([_normalize_value(v) for v in row])
        for item in sheet_spec.get("formulas", []) or []:
            cell = item.get("cell")
            formula = str(item.get("formula") or "")
            if cell and formula:
                ws[cell] = formula if formula.startswith("=") else "=" + formula
        _style_header_row(ws, 1)
        _auto_width(ws)
        if ws.max_row > 1:
            ws.freeze_panes = "A2"
        if sheet_spec.get("make_table", True) and headers and rows:
            try:
                _add_excel_table(ws, f"Table_{name}")
            except Exception:
                logger.exception("Не удалось добавить таблицу Excel")

    if created_any and default_ws.title in wb.sheetnames:
        del wb[default_ws.title]
    wb.save(output_path)


def create_excel_from_request(request_text: str) -> Tuple[str, str]:
    system_prompt = """
Ты создаешь структуру Excel-файла по просьбе пользователя. Отвечай ТОЛЬКО валидным JSON без markdown.
Верни объект:
{
  "filename": "имя_файла.xlsx",
  "message": "короткое описание файла",
  "sheets": [
    {"name":"Лист1", "headers":["Колонка 1","Колонка 2"], "rows":[["",""]], "formulas":[], "make_table": true}
  ]
}
Правила:
- Русские названия листов и столбцов.
- Если пользователь просит шаблон/трекер — создай понятные заголовки и 5-10 пустых или примерных строк.
- Не создавай больше 10 листов и 500 строк на лист.
- Формулы Excel должны начинаться с "=".
""".strip()
    user_prompt = f"Создай Excel-файл по просьбе пользователя:\n{request_text}"
    raw = _call_ai_for_excel(system_prompt, user_prompt, max_tokens=3500, temperature=0.2)
    spec = _extract_json(raw)
    file_name = _safe_filename(spec.get("filename") or "created_excel.xlsx")
    output_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4().hex[:6]}_{file_name}")
    _build_workbook_from_spec(spec, output_path)
    return output_path, spec.get("message") or "Готово, создала Excel-файл."


async def handle_excel_document(update, context) -> None:
    doc = update.message.document
    file_name = doc.file_name or "table.xlsx"
    lower_name = file_name.lower()

    if not is_supported_excel_file(file_name):
        await update.message.reply_text(
            "Пока поддерживаю только .xlsx. Если у вас .xls/.xlsm/.xlsb/.ods — откройте файл в Excel и сохраните как .xlsx, затем отправьте снова."
        )
        return

    user_request = (update.message.caption or "").strip()
    file = await context.bot.get_file(doc.file_id)
    tmp_path = None
    out_path = None

    try:
        await update.message.reply_chat_action(action="typing")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        await file.download_to_drive(tmp_path)

        user_id = get_dialog_key(update)
        # Всегда запоминаем файл. Это главное отличие: теперь можно сначала отправить Excel,
        # а затем следующим сообщением попросить его изменить.
        remember_excel_file(context, update.effective_chat.id, tmp_path, file_name, user_id)

        if not user_request:
            summary = workbook_to_text(tmp_path, max_rows_per_sheet=20, max_cols_per_sheet=12)
            add_to_conversation(user_id, "user", f"[Excel {file_name}]\n{summary[:4000]}")
            await update.message.reply_text(
                f"Excel-файл {file_name} получен и прочитан.\n"
                "Я запомнила этот файл для текущего чата. Теперь следующим сообщением напишите, что сделать, например:\n"
                "• Проанализируй просроченные задачи\n"
                "• Добавь столбец Комментарий и отметь строки, где прогресс меньше 50%\n"
                "• Создай лист Итоги по ответственным\n\n"
                "Если нужно сразу редактировать, удобнее отправлять Excel с подписью-командой."
            )
            return

        if looks_like_excel_edit_request(user_request):
            out_path, message, changes = await asyncio.to_thread(
                edit_excel_with_ai, tmp_path, file_name, user_request, user_id
            )
            if out_path:
                # Запоминаем уже измененную версию, чтобы следующие правки применялись к ней.
                remember_excel_file(context, update.effective_chat.id, out_path, os.path.basename(out_path), user_id)
                text = message
                if changes:
                    text += "\n\nЧто изменено:\n" + "\n".join(f"• {change}" for change in changes[:12])
                await update.message.reply_text(text)
                with open(out_path, "rb") as f:
                    await update.message.reply_document(document=f, filename=os.path.basename(out_path))
            else:
                await update.message.reply_text(message)
        else:
            answer = await asyncio.to_thread(analyze_excel_with_ai, tmp_path, file_name, user_request, user_id)
            await update.message.reply_text(answer)

    except ExcelProcessingError as e:
        logger.exception(f"Excel AI/parse error for {file_name}: {e}")
        await update.message.reply_text(
            "Excel-файл получен, но не удалось выполнить анализ/обработку.\n"
            f"Причина: {e}\n\n"
            "Файл не обязательно поврежден. Возможна особенность книги, лимит контекста или сбой ИИ-сервиса."
        )
    except Exception as e:
        logger.exception(f"Excel processing error for {file_name}: {e}")
        await update.message.reply_text(
            "Не удалось обработать Excel-файл. Проверьте, что это обычный .xlsx без пароля и повреждений."
        )
    finally:
        for path in (tmp_path, out_path):
            if path and os.path.exists(path):
                try:
                    os.unlink(path)
                except OSError:
                    pass


async def handle_excel_followup_text(update, context, text: str) -> bool:
    """Обрабатывает текст после ранее отправленного Excel-файла."""
    dialog_key = get_dialog_key(update)
    info = _get_recent_excel_context(context, dialog_key)
    if not info or not _is_excel_followup_request(context, text, dialog_key):
        return False

    request_text = (text or "").strip()
    if not request_text:
        return False

    user_id = dialog_key
    input_path = info["path"]
    file_name = info.get("file_name") or "table.xlsx"
    out_path = None

    try:
        await update.message.reply_chat_action(action="typing")
        if looks_like_excel_edit_request(request_text):
            await update.message.reply_text("Поняла. Вношу изменения в последний Excel-файл и пришлю новую копию.")
            out_path, message, changes = await asyncio.to_thread(
                edit_excel_with_ai, input_path, file_name, request_text, user_id
            )
            if out_path:
                remember_excel_file(context, update.effective_chat.id, out_path, os.path.basename(out_path), dialog_key)
                text_msg = message
                if changes:
                    text_msg += "\n\nЧто изменено:\n" + "\n".join(f"• {change}" for change in changes[:12])
                await update.message.reply_text(text_msg)
                with open(out_path, "rb") as f:
                    await update.message.reply_document(document=f, filename=os.path.basename(out_path))
            else:
                await update.message.reply_text(message)
        else:
            answer = await asyncio.to_thread(analyze_excel_with_ai, input_path, file_name, request_text, user_id)
            await update.message.reply_text(answer)
        context.user_data.get("awaiting_excel_request_by_dialog", {}).pop(dialog_key, None)
        return True
    except ExcelProcessingError as e:
        logger.exception(f"Excel followup AI/parse error for {file_name}: {e}")
        await update.message.reply_text(
            "Последний Excel-файл прочитан, но не удалось выполнить обработку.\n"
            f"Причина: {e}"
        )
        return True
    except Exception as e:
        logger.exception(f"Excel followup processing error for {file_name}: {e}")
        await update.message.reply_text(
            "Не удалось выполнить действие с последним Excel-файлом. Попробуйте отправить файл ещё раз с подписью-командой."
        )
        return True
    finally:
        # out_path уже скопирован в кэш через remember_excel_file, поэтому временный файл можно удалить.
        if out_path and os.path.exists(out_path):
            try:
                os.unlink(out_path)
            except OSError:
                pass


async def handle_create_excel_text(update, context, text: str) -> bool:
    if not is_create_excel_request(text):
        return False

    out_path = None
    try:
        await update.message.reply_chat_action(action="upload_document")
        out_path, message = await asyncio.to_thread(create_excel_from_request, text)
        remember_excel_file(context, update.effective_chat.id, out_path, os.path.basename(out_path), get_dialog_key(update))
        await update.message.reply_text(message)
        with open(out_path, "rb") as f:
            await update.message.reply_document(document=f, filename=os.path.basename(out_path))
        return True
    except Exception as e:
        logger.exception(f"Create Excel error: {e}")
        await update.message.reply_text("Не удалось создать Excel-файл. Попробуйте описать таблицу проще: листы, столбцы и строки.")
        return True
    finally:
        if out_path and os.path.exists(out_path):
            try:
                os.unlink(out_path)
            except OSError:
                pass
